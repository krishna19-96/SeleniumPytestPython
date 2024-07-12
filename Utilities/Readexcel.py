import openpyxl

path = r"C:\Users\Admin\PycharmProjects\Selenium\SeleniumPytestPython\testdata\Login.xlsx"


def readExcel(Sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[Sheetname]
    rows = sheet.max_row
    column = sheet.max_column
    lis2 = []
    for i in range(2, rows + 1):
        lis1 = []
        for j in range(1, column + 1):
            data = sheet.cell(row=i, column=j).value
            lis1.insert(j, data)
        lis2.insert(i, lis1)
        print(lis2)
    return lis2

liss = readExcel("LoginPage")
print(liss)
